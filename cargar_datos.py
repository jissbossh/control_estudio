import json

from pathlib import Path

import sys

import os

import re

import pandas as pd

from bs4 import BeautifulSoup

from datetime import datetime, timedelta

from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QFileDialog,
    QMessageBox,
)

from PySide6.QtCore import Qt

from PySide6.QtGui import QFont, QIcon

# ============================================================
# CARGAR CONFIGURACIONES POR DEFECTO
# ============================================================


def leer_json(ruta):
    with open(ruta, "r", encoding="utf-8") as archivo:
        return json.load(archivo)


config = leer_json("config.json")

# FUENTE_APLICACION = "Segoe UI"
FUENTE_APLICACION = config["FUENTE_APLICACION"]

# TAMANO_FUENTE = 12
TAMANO_FUENTE = int(config["TAMANO_FUENTE"])

TEMA = config["TEMA"]

# ============================================================
# COLUMNAS
# ============================================================

COLUMNAS_ACTIVIDADES = [
    "Momento",
    "RAC",
    "Nombre actividad",
    "Descripción",
    "Tipo de actividad",
    "Peso evaluativo",
    "Actividad inicia en",
    "Actividad finaliza en",
    "Alerta de cierre en",
    "Fecha entrega realimentación",
]


COLUMNAS_RESUMEN = [
    "PERIODO",
    "MOMENTO",
    "RAC",
    "CURSO",
    "TAREA",
    "ESTADO",
    "FECHA INDICADA",
    "FECHA INICIO",
    "FECHA ENTREGA",
    "FECHA CALIFICADA",
    "FALTAN",
    "CALIFICACION",
]


# ============================================================
# NORMALIZAR ROWSPAN Y COLSPAN
# ============================================================


def normalizar_tabla(tabla):

    filas_html = tabla.find_all("tr")

    matriz = []

    ocupadas = {}

    for fila_idx, tr in enumerate(filas_html):

        while len(matriz) <= fila_idx:
            matriz.append([])

        columna_idx = 0

        celdas = tr.find_all(["td", "th"], recursive=False)

        for celda in celdas:

            while (fila_idx, columna_idx) in ocupadas:
                columna_idx += 1

            texto = celda.get_text(" ", strip=True)

            # ------------------------------------------------
            # ROWSPAN
            # ------------------------------------------------

            try:

                rowspan = int(celda.get("rowspan", 1))

            except (ValueError, TypeError):

                rowspan = 1

            # ------------------------------------------------
            # COLSPAN
            # ------------------------------------------------

            try:

                colspan = int(celda.get("colspan", 1))

            except (ValueError, TypeError):

                colspan = 1

            # ------------------------------------------------
            # COLOCAR CELDA
            # ------------------------------------------------

            for r in range(rowspan):

                for c in range(colspan):

                    fila_destino = fila_idx + r

                    columna_destino = columna_idx + c

                    while len(matriz) <= fila_destino:
                        matriz.append([])

                    while len(matriz[fila_destino]) <= columna_destino:

                        matriz[fila_destino].append("")

                    matriz[fila_destino][columna_destino] = texto

                    if r > 0:

                        ocupadas[
                            (
                                fila_destino,
                                columna_destino,
                            )
                        ] = True

            columna_idx += colspan

    return matriz


# ============================================================
# OBTENER INFORMACIÓN
# ============================================================


def obtener_informacion_general(soup):

    datos = []

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    titulo = soup.find("title")

    if titulo:

        datos.append(
            {
                "Etiqueta": "title",
                "Texto": titulo.get_text(" ", strip=True),
            }
        )

    # --------------------------------------------------------
    # H1, H2, H3
    # --------------------------------------------------------

    for etiqueta in ["h1", "h2", "h3"]:

        for elemento in soup.find_all(etiqueta):

            texto = elemento.get_text(" ", strip=True)

            if texto:

                datos.append(
                    {
                        "Etiqueta": etiqueta,
                        "Texto": texto,
                    }
                )

    # --------------------------------------------------------
    # P
    # --------------------------------------------------------

    for p in soup.find_all("p"):

        texto = p.get_text(" ", strip=True)

        if texto:

            datos.append(
                {
                    "Etiqueta": "p",
                    "Texto": texto,
                }
            )

    return datos


# ============================================================
# OBTENER CURSO
# ============================================================


def obtener_curso_desde_informacion(df_informacion):

    for _, fila in df_informacion.iterrows():

        etiqueta = str(fila["Etiqueta"]).strip().lower()

        texto = str(fila["Texto"]).strip()

        if etiqueta == "p":

            if "Curso Teórico" in texto:

                curso = texto.split(" - Curso Teórico")[0].strip()

                return curso

    return ""


# ============================================================
# OBTENER PERIODO
# ============================================================


def obtener_periodo_desde_informacion(df_informacion):

    for _, fila in df_informacion.iterrows():

        etiqueta = str(fila["Etiqueta"]).strip().lower()

        texto = str(fila["Texto"]).strip()

        if etiqueta != "p":
            continue

        if "PERIODO" not in texto.upper():
            continue

        patron = r"\b(\d{4})\b" r"\s+" r".*?" r"\bPERIODO\b" r"\s+" r"(\d{2}-\d{2})"

        resultado = re.search(patron, texto, re.IGNORECASE)

        if resultado:

            anio = resultado.group(1)

            periodo = resultado.group(2)

            return f"{anio}-{periodo}"

    return ""


# ============================================================
# OBTENER ACTIVIDADES
# ============================================================


def obtener_actividades(soup):

    tablas = soup.find_all("table")

    if not tablas:

        return pd.DataFrame(columns=COLUMNAS_ACTIVIDADES)

    matriz = normalizar_tabla(tablas[0])

    actividades = []

    for fila in matriz:

        fila = [str(valor).strip() if valor is not None else "" for valor in fila]

        if not any(fila):
            continue

        # ----------------------------------------------------
        # IGNORAR ENCABEZADO
        # ----------------------------------------------------

        if fila[0] == ("1. Momento de la e-evaluación"):

            continue

        # ----------------------------------------------------
        # IGNORAR TÍTULOS DE SECCIÓN
        # ----------------------------------------------------

        texto_fila = " ".join(fila)

        if "Actividades correspondientes al" in texto_fila:

            continue

        # ----------------------------------------------------
        # ASEGURAR 10 COLUMNAS
        # ----------------------------------------------------

        if len(fila) < 10:

            fila += [""] * (10 - len(fila))

        elif len(fila) > 10:

            fila = fila[:10]

        actividades.append(fila)

    return pd.DataFrame(actividades, columns=COLUMNAS_ACTIVIDADES)


# ============================================================
# CONVERTIR FECHA
# ============================================================


def convertir_fecha(fecha):

    if not fecha:
        return ""

    meses = {
        "ENE": "01",
        "FEB": "02",
        "MAR": "03",
        "ABR": "04",
        "MAY": "05",
        "JUN": "06",
        "JUL": "07",
        "AGO": "08",
        "SEP": "09",
        "OCT": "10",
        "NOV": "11",
        "DIC": "12",
    }

    fecha = fecha.strip()

    patron = r"(\d{1,2})/" r"([A-ZÁÉÍÓÚÑ]{3})/" r"(\d{4})"

    resultado = re.search(patron, fecha.upper())

    if not resultado:

        return fecha

    dia = resultado.group(1).zfill(2)

    mes = resultado.group(2)

    anio = resultado.group(3)

    mes = meses.get(mes, mes)

    return f"{dia}/{mes}/{anio}"


# ============================================================
# FECHA A DATETIME
# ============================================================


def fecha_a_datetime(fecha):

    if not fecha:
        return None

    try:

        return datetime.strptime(fecha, "%d/%m/%Y")

    except ValueError:

        return None


# ============================================================
# CONVERTIR FECHA A DATETIME CON HORA
# ============================================================


def convertir_fecha_hora(fecha, hora="00:00:00"):

    if not fecha:

        return None

    fecha = str(fecha).strip()

    # --------------------------------------------------------
    # ISO
    # --------------------------------------------------------

    try:

        return datetime.fromisoformat(fecha.replace("Z", ""))

    except ValueError:

        pass

    # --------------------------------------------------------
    # DD/MM/YYYY
    # --------------------------------------------------------

    try:

        fecha_objeto = datetime.strptime(fecha, "%d/%m/%Y")

        hora_objeto = datetime.strptime(hora, "%H:%M:%S").time()

        return datetime.combine(fecha_objeto.date(), hora_objeto)

    except ValueError:

        return None


# ============================================================
# FECHA A ISO
# ============================================================


def fecha_a_iso(fecha, hora="00:00:00"):

    fecha_objeto = convertir_fecha_hora(fecha, hora)

    if fecha_objeto is None:

        return ""

    return fecha_objeto.strftime("%Y-%m-%dT%H:%M:%S")


# ============================================================
# FECHA INDICADA
# ============================================================


def calcular_fecha_indicada(fecha_entrega):

    fecha = fecha_a_datetime(fecha_entrega)

    if fecha is None:

        return ""

    fecha_indicada = fecha - timedelta(days=5)

    return fecha_indicada.strftime("%d/%m/%Y")


# ============================================================
# FECHA CALIFICADA
# ============================================================


def obtener_fecha_realimentacion(fecha):

    if not fecha:

        return ""

    fecha = fecha.strip()

    partes = fecha.split("-")

    if len(partes) >= 2:

        segunda_fecha = partes[-1].strip()

        return convertir_fecha(segunda_fecha)

    return convertir_fecha(fecha)


# ============================================================
# CALCULAR FALTAN
# ============================================================


def calcular_faltan(fecha_entrega):

    if not fecha_entrega:

        return ""

    fecha_objetivo = fecha_a_datetime(fecha_entrega)

    if fecha_objetivo is None:

        return ""

    # La actividad termina a las 23:55

    fecha_objetivo = fecha_objetivo.replace(hour=23, minute=55, second=0)

    ahora = datetime.now()

    diferencia = fecha_objetivo - ahora

    if diferencia.total_seconds() <= 0:

        return "VENCIDA"

    dias = diferencia.days

    segundos = diferencia.seconds

    horas = segundos // 3600

    minutos = (segundos % 3600) // 60

    return f"Faltan {dias} días, " f"{horas} horas, " f"{minutos} minutos"


# ============================================================
# CREAR RESUMEN
# ============================================================


def crear_resumen(df_actividades, periodo, curso):

    filas = []

    for _, actividad in df_actividades.iterrows():

        # ----------------------------------------------------
        # MOMENTO
        # ----------------------------------------------------

        momento = str(actividad["Momento"]).strip()

        # ----------------------------------------------------
        # RAC
        # ----------------------------------------------------

        rac = str(actividad["RAC"]).strip()

        # ----------------------------------------------------
        # TAREA
        # ----------------------------------------------------

        tarea = str(actividad["Nombre actividad"]).strip()

        # ----------------------------------------------------
        # FECHA INICIO
        # ----------------------------------------------------

        fecha_inicio_original = convertir_fecha(str(actividad["Actividad inicia en"]))

        fecha_inicio = fecha_a_iso(fecha_inicio_original, "00:00:00")

        # ----------------------------------------------------
        # FECHA ENTREGA
        # ----------------------------------------------------

        fecha_entrega_original = convertir_fecha(
            str(actividad["Actividad finaliza en"])
        )

        fecha_entrega = fecha_a_iso(fecha_entrega_original, "23:55:00")

        # ----------------------------------------------------
        # FECHA INDICADA
        # ----------------------------------------------------

        fecha_indicada_original = calcular_fecha_indicada(fecha_entrega_original)

        fecha_indicada = fecha_a_iso(fecha_indicada_original, "23:55:00")

        # ----------------------------------------------------
        # FECHA CALIFICADA
        # ----------------------------------------------------

        fecha_calificada_original = obtener_fecha_realimentacion(
            str(actividad["Fecha entrega realimentación"])
        )

        fecha_calificada = fecha_a_iso(fecha_calificada_original, "00:00:00")

        # ----------------------------------------------------
        # FALTAN
        # ----------------------------------------------------

        faltan = calcular_faltan(fecha_entrega_original)

        # ----------------------------------------------------
        # FILA
        # ----------------------------------------------------

        fila = {
            "PERIODO": periodo,
            "MOMENTO": momento.upper(),
            "RAC": rac,
            "CURSO": curso,
            "TAREA": tarea,
            "ESTADO": "PENDIENTE",
            "FECHA INDICADA": fecha_indicada,
            "FECHA INICIO": fecha_inicio,
            "FECHA ENTREGA": fecha_entrega,
            "FECHA CALIFICADA": (fecha_calificada),
            "FALTAN": faltan,
            "CALIFICACION": 0,
        }

        filas.append(fila)

    return pd.DataFrame(filas, columns=COLUMNAS_RESUMEN)


# ============================================================
# CONVERTIR COLUMNAS DE FECHA
# ============================================================


def convertir_columnas_fecha_excel(df):

    columnas_fecha = [
        "FECHA INDICADA",
        "FECHA INICIO",
        "FECHA ENTREGA",
        "FECHA CALIFICADA",
    ]

    for columna in columnas_fecha:

        if columna not in df.columns:

            continue

        df[columna] = pd.to_datetime(
            df[columna], format="%Y-%m-%dT%H:%M:%S", errors="coerce"
        )

    return df


# ============================================================
# LIMPIAR NOMBRE DE HOJA
# ============================================================


def limpiar_nombre_hoja(nombre):

    if not nombre:

        return "Resumen"

    nombre = re.sub(r"[:\\/?*\[\]]", "", nombre)

    nombre = nombre.strip()

    if not nombre:

        nombre = "Resumen"

    nombre = nombre[:31]

    return nombre


# ============================================================
# FORMATEAR EXCEL
# ============================================================


def ajustar_excel(archivo):

    from openpyxl import load_workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    libro = load_workbook(archivo)

    color_encabezado = "2E75B6"

    formato_fecha_hora = "yyyy-mm-dd\\Thh:mm:ss"

    columnas_fecha = [
        "FECHA INDICADA",
        "FECHA INICIO",
        "FECHA ENTREGA",
        "FECHA CALIFICADA",
    ]

    for hoja in libro.worksheets:

        # ----------------------------------------------------
        # CONGELAR ENCABEZADO
        # ----------------------------------------------------

        hoja.freeze_panes = "A2"

        # ----------------------------------------------------
        # FILTROS
        # ----------------------------------------------------

        if hoja.max_row > 1:

            hoja.auto_filter.ref = hoja.dimensions

        # ----------------------------------------------------
        # ENCABEZADOS
        # ----------------------------------------------------

        for celda in hoja[1]:

            celda.font = Font(bold=True, color="FFFFFF")

            celda.fill = PatternFill(fill_type="solid", fgColor=color_encabezado)

            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # ----------------------------------------------------
        # CONTENIDO
        # ----------------------------------------------------

        for fila in hoja.iter_rows():

            for celda in fila:

                celda.alignment = Alignment(vertical="top", wrap_text=True)

        # ----------------------------------------------------
        # FORMATO FECHA-HORA
        # ----------------------------------------------------

        encabezados = {}

        for celda in hoja[1]:

            encabezados[str(celda.value).strip()] = celda.column

        for nombre_columna in columnas_fecha:

            if nombre_columna not in encabezados:

                continue

            numero_columna = encabezados[nombre_columna]

            for fila in range(2, hoja.max_row + 1):

                celda = hoja.cell(row=fila, column=numero_columna)

                if celda.value is not None:

                    celda.number_format = formato_fecha_hora

        # ----------------------------------------------------
        # ANCHO COLUMNAS
        # ----------------------------------------------------

        for columna in hoja.columns:

            letra = columna[0].column_letter

            maximo = 0

            for celda in columna:

                if celda.value is not None:

                    longitud = len(str(celda.value))

                    maximo = max(maximo, longitud)

            ancho = min(max(maximo + 2, 12), 60)

            hoja.column_dimensions[letra].width = ancho

    libro.save(archivo)


# ============================================================
# MOSTRAR ERROR
# ============================================================


def mostrar_error(parent, titulo, mensaje):

    QMessageBox.critical(parent, titulo, mensaje)


# ============================================================
# MOSTRAR ADVERTENCIA
# ============================================================


def mostrar_advertencia(parent, titulo, mensaje):

    QMessageBox.warning(parent, titulo, mensaje)


# ========================================================
# TEMA CLARO
# ========================================================


def aplicar_tema_claro(ventana):

    ventana.setStyleSheet(f"""
            QWidget {{
                font-family: "{FUENTE_APLICACION}";
                font-size: {TAMANO_FUENTE}pt;
                color: #212529;
            }}

            QMainWindow,
            QDialog {{
                background-color: #f5f7fa;
            }}

            QLabel {{
                color: #212529;
            }}

            QLineEdit,
            QComboBox,
            QDateTimeEdit,
            QDoubleSpinBox {{
                padding: 6px;
                border: 1px solid #ced4da;
                border-radius: 5px;
                background-color: white;
                color: #212529;
            }}

            QPushButton {{
                padding: 7px 14px;
                border: 1px solid #ced4da;
                border-radius: 5px;
                background-color: white;
                color: #212529;
            }}

            QPushButton:hover {{
                background-color: #e9ecef;
            }}

            QTableWidget {{
                background-color: white;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
                color: #212529;
                selection-background-color: #cfe2ff;
                selection-color: #000000;
            }}

            QHeaderView::section {{
                background-color: #212529;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }}

            QComboBox QAbstractItemView {{
                background-color: white;
                color: #212529;
                selection-background-color: #cfe2ff;
            }}

            QDialogButtonBox QPushButton {{
                min-width: 90px;
            }}
            """)


# ========================================================
# TEMA OSCURO
# ========================================================


def aplicar_tema_oscuro(ventana):

    ventana.setStyleSheet(f"""
            QWidget {{
                font-family: "{FUENTE_APLICACION}";
                font-size: {TAMANO_FUENTE}pt;
                background-color: #121212;
                color: #eeeeee;
            }}

            QMainWindow,
            QDialog {{
                background-color: #121212;
            }}

            QLabel {{
                color: #eeeeee;
            }}

            QLineEdit,
            QComboBox,
            QDateTimeEdit,
            QDoubleSpinBox {{
                padding: 6px;
                border: 1px solid #444444;
                border-radius: 5px;
                background-color: #1e1e1e;
                color: #eeeeee;
            }}

            QPushButton {{
                padding: 7px 14px;
                border: 1px solid #444444;
                border-radius: 5px;
                background-color: #242424;
                color: #eeeeee;
            }}

            QPushButton:hover {{
                background-color: #333333;
            }}

            QTableWidget {{
                background-color: #1e1e1e;
                alternate-background-color: #252525;
                gridline-color: #3a3a3a;
                color: #eeeeee;
                selection-background-color: #264f78;
                selection-color: white;
            }}

            QHeaderView::section {{
                background-color: #000000;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }}

            QComboBox QAbstractItemView {{
                background-color: #1e1e1e;
                color: white;
                selection-background-color: #264f78;
            }}

            QScrollBar:vertical {{
                background: #1e1e1e;
                width: 12px;
            }}

            QScrollBar::handle:vertical {{
                background: #555555;
                border-radius: 5px;
            }}

            QDialogButtonBox QPushButton {{
                min-width: 90px;
            }}

            QMessageBox {{
                background-color: #1e1e1e;
                color: #eeeeee;
            }}

            QMessageBox QLabel {{
                color: #eeeeee;
            }}

            QMessageBox QPushButton {{
                background-color: #333333;
                color: white;
                border: 1px solid #555555;
            }}

            QMessageBox QPushButton:hover {{
                background-color: #444444;
            }}

            QTableWidget {{
            background-color: #1e1e1e;
            color: white;
            gridline-color: #333333;
            }}

            QHeaderView::section {{
            background-color: #000000;
            color: white;
            padding: 5px;
            border: 1px solid #333333;
            }}

            QTableCornerButton::section {{
            background-color: #000000;
            border: 1px solid #333333;
            }}
            """)


# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================


def main():

    # ========================================================
    # CREAR APLICACIÓN QT
    # ========================================================

    app = QApplication.instance()

    if app is None:

        app = QApplication(sys.argv)

    # ========================================================
    # VENTANA PADRE
    # ========================================================

    parent = QWidget()

    if TEMA == "CLARO":
        aplicar_tema_claro(parent)
    else:
        aplicar_tema_oscuro(parent)

    parent.setWindowTitle("Procesador de agenda")

    ruta_icono = Path(__file__).resolve().parent / "icono.png"

    if ruta_icono.exists():

        parent.setWindowIcon(QIcon(str(ruta_icono)))
    # No mostramos la ventana.
    # Solo se utiliza como padre de los diálogos.

    parent.hide()

    # ========================================================
    # SELECCIONAR HTML
    # ========================================================

    archivo_html, _ = QFileDialog.getOpenFileName(
        parent,
        "Selecciona la agenda HTML",
        "",
        "Archivos HTML (*.html *.htm);;" "Todos los archivos (*)",
    )

    if not archivo_html:

        return

    # ========================================================
    # LEER HTML
    # ========================================================

    try:

        with open(archivo_html, "r", encoding="cp1252", errors="replace") as archivo:

            html = archivo.read()

    except Exception as error:

        mostrar_error(
            parent, "Error", "No se pudo leer el archivo HTML.\n\n" f"{error}"
        )

        return

    # ========================================================
    # BEAUTIFULSOUP
    # ========================================================

    try:

        soup = BeautifulSoup(html, "html.parser")

    except Exception as error:

        mostrar_error(parent, "Error", "No se pudo procesar el HTML.\n\n" f"{error}")

        return

    # ========================================================
    # INFORMACIÓN
    # ========================================================

    try:

        informacion = obtener_informacion_general(soup)

        df_informacion = pd.DataFrame(informacion, columns=["Etiqueta", "Texto"])

    except Exception as error:

        mostrar_error(
            parent, "Error", "No se pudo obtener la información.\n\n" f"{error}"
        )

        return

    # ========================================================
    # CURSO
    # ========================================================

    curso = obtener_curso_desde_informacion(df_informacion)

    # ========================================================
    # PERIODO
    # ========================================================

    periodo = obtener_periodo_desde_informacion(df_informacion)

    # ========================================================
    # ACTIVIDADES
    # ========================================================

    try:

        actividades = obtener_actividades(soup)

    except Exception as error:

        mostrar_error(
            parent, "Error", "No se pudieron obtener " "las actividades.\n\n" f"{error}"
        )

        return

    # ========================================================
    # CREAR RESUMEN
    # ========================================================

    resumen = crear_resumen(actividades, periodo, curso)

    # ========================================================
    # CONVERTIR FECHAS
    # ========================================================

    resumen = convertir_columnas_fecha_excel(resumen)

    # ========================================================
    # COMPROBAR CURSO
    # ========================================================

    if not curso:

        mostrar_advertencia(
            parent,
            "Curso no encontrado",
            "No fue posible obtener el nombre "
            "del curso desde la información "
            "del HTML.\n\n"
            "Se utilizará 'Resumen' como nombre "
            "de la hoja.",
        )

    # ========================================================
    # NOMBRE DE HOJA
    # ========================================================

    nombre_hoja = limpiar_nombre_hoja(curso)

    # ========================================================
    # SELECCIONAR ARCHIVO DE SALIDA
    # ========================================================

    nombre_original = os.path.splitext(os.path.basename(archivo_html))[0]

    archivo_excel, _ = QFileDialog.getSaveFileName(
        parent,
        "Guardar resultado",
        os.path.join(os.path.dirname(archivo_html), f"{nombre_original}_resumen.xlsx"),
        "Archivo Excel (*.xlsx)",
    )

    if not archivo_excel:

        return

    # ========================================================
    # ASEGURAR EXTENSIÓN
    # ========================================================

    if not archivo_excel.lower().endswith(".xlsx"):

        archivo_excel += ".xlsx"

    # ========================================================
    # SI EXISTE, ELIMINAR
    # ========================================================

    if os.path.exists(archivo_excel):

        try:

            os.remove(archivo_excel)

        except PermissionError:

            mostrar_error(
                parent,
                "Archivo abierto",
                "El archivo Excel está abierto.\n\n" "Ciérralo y vuelve a intentarlo.",
            )

            return

        except Exception as error:

            mostrar_error(
                parent, "Error", "No se pudo reemplazar " "el archivo.\n\n" f"{error}"
            )

            return

    # ========================================================
    # GUARDAR SOLO RESUMEN
    # ========================================================

    try:

        with pd.ExcelWriter(archivo_excel, engine="openpyxl") as writer:

            resumen.to_excel(writer, sheet_name=nombre_hoja, index=False)

        # ----------------------------------------------------
        # FORMATEAR EXCEL
        # ----------------------------------------------------

        ajustar_excel(archivo_excel)

    except PermissionError:

        mostrar_error(
            parent,
            "Error al guardar",
            "No se pudo guardar el archivo.\n\n"
            "Comprueba que el archivo Excel "
            "no esté abierto.",
        )

        return

    except Exception as error:

        mostrar_error(
            parent,
            "Error al guardar",
            "Ocurrió un error al guardar " "el Excel.\n\n" f"{error}",
        )

        return

    # ========================================================
    # MENSAJE FINAL
    # ========================================================

    QMessageBox.information(
        parent,
        "Proceso terminado",
        "Datos extraídos correctamente.\n\n"
        f"Curso: {curso}\n"
        f"Periodo: {periodo}\n"
        f"Actividades: {len(actividades)}\n\n"
        f"Hoja creada: {nombre_hoja}\n\n"
        f"Archivo generado:\n"
        f"{archivo_excel}",
    )


# ============================================================
# EJECUTAR
# ============================================================

if __name__ == "__main__":

    main()
